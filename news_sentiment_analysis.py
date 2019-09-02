from urllib.request import urlopen
from bs4 import BeautifulSoup  # pip install BeautifulSoup4
import openpyxl  # pip install openpyxl   for excel
import jieba  # pip install jieba
import pandas as pd  # pip install pandas
from snownlp import SnowNLP  # pip install snownlp
# pip install xlrd
# ##### for schedule #####
import requests  # pip install requests
import schedule  # pip install schedule
import time

def job():
    url = "https://www.jin10.com/"
    response = urlopen(url)
    html = BeautifulSoup(response)
    # print(html)
    jieba.set_dictionary("jieba_dict/dict.txt.big")  # 設定繁體中文詞庫
    jieba.load_userdict("jieba_dict/myself_def_dict.txt")  # 設定自訂詞庫
    with open("jieba_dict/stopwords.txt", "r", encoding="utf-8-sig") as f:  # 設定停用字
        stops = f.read().split("\n")  # 讀取停用詞並存於stops串列中

    with open("jieba_dict/American.txt", "r", encoding="utf-8-sig") as g:  # 設定美國關鍵字
        American = g.read().split("\n")  # 讀取美國關鍵字並存於American串列中
    # print(set(American))

    with open("jieba_dict/Europe.txt", "r", encoding="utf-8-sig") as h:  # 設定歐洲關鍵字
        Europe = h.read().split("\n")  # 讀取歐洲關鍵字並存於Europe串列中
    # print(set(Europe))

    workbook_1 = openpyxl.Workbook()  # 利用 Workbook 方法建立一個新的工作簿
    sheet_1 = workbook_1.worksheets[0]  # 取得工作表  # 取得工作簿的第一個工作表
    listtitle_1 = ["comments", "date"]
    sheet_1.append(listtitle_1)  # 以 append 方法加入串列資料

    workbook_2 = openpyxl.Workbook()  # 利用 Workbook 方法建立一個新的工作簿
    sheet_2 = workbook_2.worksheets[0]  # 取得工作表  # 取得工作簿的第一個工作表
    listtitle_2 = ["comments", "date"]
    sheet_2.append(listtitle_1)  # 以 append 方法加入串列資料

    workbook_3 = openpyxl.Workbook()  # 利用 Workbook 方法建立一個新的工作簿
    sheet_3 = workbook_3.worksheets[0]  # 取得工作表  # 取得工作簿的第一個工作表
    listtitle_3 = ["comments", "date"]
    sheet_3.append(listtitle_3)  # 以 append 方法加入串列資料

    workbook_4 = openpyxl.Workbook()  # 利用 Workbook 方法建立一個新的工作簿
    sheet_4 = workbook_4.worksheets[0]  # 取得工作表  # 取得工作簿的第一個工作表
    listtitle_4 = ["comments", "date"]
    sheet_4.append(listtitle_4)  # 以 append 方法加入串列資料

    workbook = openpyxl.Workbook()  # 利用 Workbook 方法建立一個新的工作簿
    sheet = workbook.worksheets[0]  # 取得工作表  # 取得工作簿的第一個工作表

    # news = html.find_all("div", class_="jin-flash")
    news = html.find_all("div", class_="jin-flash_item")
    # print(news)
    listtitle = ["comments", "date"]
    sheet.append(listtitle)  # 以 append 方法加入串列資料

    ticks = time.strftime("%Y-%m-%d-%H", time.localtime())  # 紀錄時間檔案
    print("當前時間為:", ticks)
    for r in news:
        en = r.find("div", class_="jin-flash_time")
        fn = r.find("div", class_="jin-flash_b")
        #print(fn.text, en.text)
        #print(len(fn.text))
        # ############## jieba function start ###############
        sentence = fn.text
        breakword = jieba.cut(sentence, cut_all=False)  # 設定精確模式
        list_breakword = list(breakword)
        # print("精確模式\t" + " | ".join(breakword))
        # ############## jieba function end ###############
        # ############## classify function start ###############
        if len(set(American) & set(list_breakword)) >= 1 and len(set(Europe) & set(list_breakword)) >= 1:
            # print("歐洲+美國\t" + fn.text)
            listdata_1 = [fn.text, en.text]
            sheet_1.append(listdata_1)  # 以 append 方法加入串列資料
        elif len(set(American) & set(list_breakword)) >= 1:
            # print("美國\t" + fn.text)
            listdata_2 = [fn.text, en.text]
            sheet_2.append(listdata_2)  # 以 append 方法加入串列資料
        elif len(set(Europe) & set(list_breakword)) >= 1:
            # print("歐洲\t" + fn.text)
            listdata_3 = [fn.text, en.text]
            sheet_3.append(listdata_3)  # 以 append 方法加入串列資料
        elif len(fn.text) > 0:
            # print("other\t" + fn.text)
            # print(len(fn.text))
            listdata_4 = [fn.text, en.text]
            sheet_4.append(listdata_4)  # 以 append 方法加入串列資料
        # ############## classify function end ###############

    workbook_1.save('comments_data/' + str(ticks) + '_american_europe.xlsx')
    workbook_2.save('comments_data/' + str(ticks) + '_american.xlsx')
    workbook_3.save('comments_data/' + str(ticks) + '_europe.xlsx')
    workbook_4.save('comments_data/' + str(ticks) + '_other.xlsx')
    # ############## ############### ###############


    def get_sentiment_cn(text):
        s = SnowNLP(text)
        return s.sentiments  # s.sentiments 經語意分析後得到數據

    df_American = pd.read_excel('comments_data/' + str(ticks) + '_american.xlsx')  # 讀取 excel 檔案
    df_Europe = pd.read_excel('comments_data/' + str(ticks) + '_europe.xlsx')

    # text = df.comments.iloc[0]  # 選取特定列

    df_American["sentiment"] = df_American.comments.apply(get_sentiment_cn)  # 將各個欄位做語意分析並寫入 df["sentiment"]
    df_American.sort_values((['sentiment'])[:1])  # 以情感分數最為排序

    df_Europe["sentiment"] = df_Europe.comments.apply(get_sentiment_cn)  # 將各個欄位做語意分析並寫入 df["sentiment"]
    df_Europe.sort_values((['sentiment'])[:1])  # 以情感分數最為排序

    # print(df_American["sentiment"])
    # print(df_American.head())
    # print(df_American.sentiment.mean())
    # #################### Adjustment score and weight start ####################

    Amer_sent_median = (df_American.sentiment.median() - 0.5)*5
    Euro_sent_median = (df_Europe.sentiment.median() - 0.5)*5
    score_difference = Euro_sent_median - Amer_sent_median
    # nan is never equal to nan
    if Amer_sent_median != Amer_sent_median:
        Amer_sent_median = 0
    if Euro_sent_median != Euro_sent_median:
        Euro_sent_median = 0

    # #################### Adjustment score and weight end ####################

    print("American sentiment score : ", Amer_sent_median)
    print("Europe sentiment score : ", Euro_sent_median)
    print("score_difference : ", score_difference)

    workbook_5 = openpyxl.Workbook()  # 利用 Workbook 方法建立一個新的工作簿
    sheet_5 = workbook_5.worksheets[0]  # 取得工作表  # 取得工作簿的第一個工作表
    listtitle_5 = ["American sentiment score : ", Amer_sent_median]
    sheet_5.append(listtitle_5)
    listtitle_5 = ["Europe sentiment score : ", Euro_sent_median]
    sheet_5.append(listtitle_5)
    listtitle_5 = ["score_difference : ", score_difference]
    sheet_5.append(listtitle_5)
    workbook_5.save('sentiment_score/' + str(ticks) + '_score_difference.xlsx')
    # ####################  ####################
    # print(df_American.sort_values((['sentiment'])[:1]))
    df_American.sort_values((['sentiment'])[:1]).to_excel('sentiment_data/' + str(ticks) + '_American_sens.xlsx',
                                                          encoding="utf-8-sig", index=False)  # 將結果寫入excel, 不印index
    df_Europe.sort_values((['sentiment'])[:1]).to_excel('sentiment_data/' + str(ticks) + '_Europe_sens.xlsx',
                                                        encoding="utf-8-sig", index=False)  # 將結果寫入excel, 不印index


schedule.every(30).minutes.do(job)
#schedule.every().day.at("01:00").do(job)
#schedule.every().day.at("05:00").do(job)
#schedule.every().day.at("09:00").do(job)
#schedule.every().day.at("13:00").do(job)
#schedule.every().day.at("17:00").do(job)
#schedule.every().day.at("21:00").do(job)
while True:
    schedule.run_pending()
    time.sleep(3)
